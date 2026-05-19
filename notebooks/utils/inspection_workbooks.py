from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")
UNMAPPED_FILL = PatternFill("solid", fgColor="F4CCCC")
USD_NUMBER_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
GENERIC_NUMBER_FORMAT = '#,##0.00;[Red]-#,##0.00'
PERCENT_DISPLAY_FORMAT = '0.0"%"'


@dataclass(frozen=True)
class CsvSheetSpec:
    """Describe one CSV-backed sheet inside an inspection workbook."""

    sheet_name: str
    csv_path: Path


@dataclass(frozen=True)
class InspectionWorkbookConfig:
    """Formatting rules that can be reused across inspection workbooks."""

    financial_metric_columns: frozenset[str] = frozenset()
    explicit_currency_columns: frozenset[str] = frozenset({"value_usd_millions"})
    explicit_percentage_columns: frozenset[str] = frozenset()
    currency_row_allowed_values_by_column: dict[str, frozenset[str]] | None = None
    integer_label_columns: frozenset[str] = frozenset({"fiscal_year"})
    text_label_columns: frozenset[str] = frozenset(
        {
            "fiscal_quarter",
            "period",
            "period_type",
            "statement_or_tab",
            "source_file",
            "source_platform",
            "retrieval_date",
            "unit_scale",
            "original_line_item",
            "canonical_metric",
        }
    )
    percentage_row_label_column: str | None = "original_line_item"
    percentage_row_label_tokens: tuple[str, ...] = ("margin", "rate", "yield", "growth", "percent")
    unmapped_metric_column: str | None = "canonical_metric"
    unmapped_metric_value: str = "UNMAPPED"
    max_column_width: int = 40


@dataclass(frozen=True)
class InspectionWorkbookResult:
    """Report what was actually included in a generated inspection workbook."""

    workbook_path: Path
    included_sheets: tuple[str, ...]
    csv_files_used: tuple[Path, ...]
    missing_csv_files: tuple[Path, ...]


def is_missing_display_value(value: object) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def is_percentage_like_label(value: object, tokens: tuple[str, ...]) -> bool:
    text = str(value).lower()
    return any(token in text for token in tokens)


def _validate_sheet_name(sheet_name: str) -> None:
    if len(sheet_name) > 31:
        raise ValueError(f"Excel sheet name exceeds 31 characters: {sheet_name}")


def _write_dataframe_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    config: InspectionWorkbookConfig,
) -> None:
    _validate_sheet_name(sheet_name)
    worksheet = workbook.create_sheet(title=sheet_name)

    for column_index, column_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    unmapped_column_index = (
        df.columns.get_loc(config.unmapped_metric_column) + 1
        if config.unmapped_metric_column and config.unmapped_metric_column in df.columns
        else None
    )
    percentage_label_column_index = (
        df.columns.get_loc(config.percentage_row_label_column) + 1
        if config.percentage_row_label_column and config.percentage_row_label_column in df.columns
        else None
    )
    currency_rule_column_indices = {
        column_name: df.columns.get_loc(column_name) + 1
        for column_name in (config.currency_row_allowed_values_by_column or {})
        if column_name in df.columns
    }

    for row_index, row in enumerate(df.itertuples(index=False, name=None), start=2):
        row_has_unmapped_metric = (
            unmapped_column_index is not None
            and row[unmapped_column_index - 1] == config.unmapped_metric_value
        )
        percentage_row_label = (
            row[percentage_label_column_index - 1]
            if percentage_label_column_index is not None
            else ""
        )
        row_allows_currency = True
        if config.currency_row_allowed_values_by_column:
            row_allows_currency = all(
                row[currency_rule_column_indices[column_name] - 1] in allowed_values
                for column_name, allowed_values in config.currency_row_allowed_values_by_column.items()
                if column_name in currency_rule_column_indices
            )

        for column_index, value in enumerate(row, start=1):
            column_name = df.columns[column_index - 1]
            cell_value = None if is_missing_display_value(value) else value
            if cell_value is not None and column_name in config.text_label_columns:
                cell_value = str(cell_value)
            cell = worksheet.cell(row=row_index, column=column_index, value=cell_value)

            if row_has_unmapped_metric:
                cell.fill = UNMAPPED_FILL
            elif cell_value is None:
                cell.fill = MISSING_FILL

            if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool) and not pd.isna(cell_value):
                if column_name in config.integer_label_columns:
                    cell.number_format = "0"
                elif column_name in config.explicit_currency_columns:
                    if is_percentage_like_label(
                        percentage_row_label,
                        config.percentage_row_label_tokens,
                    ):
                        cell.number_format = PERCENT_DISPLAY_FORMAT
                    elif row_allows_currency:
                        cell.number_format = USD_NUMBER_FORMAT
                    else:
                        cell.number_format = GENERIC_NUMBER_FORMAT
                elif column_name in config.explicit_percentage_columns:
                    cell.number_format = PERCENT_DISPLAY_FORMAT
                elif is_percentage_like_label(column_name, config.percentage_row_label_tokens):
                    cell.number_format = PERCENT_DISPLAY_FORMAT
                elif column_name in config.financial_metric_columns:
                    cell.number_format = USD_NUMBER_FORMAT
                else:
                    cell.number_format = GENERIC_NUMBER_FORMAT
            elif cell_value is not None and column_name in config.text_label_columns:
                cell.number_format = "@"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(df.columns, start=1):
        column_values = [str(column_name)]
        column_values.extend("" if is_missing_display_value(value) else str(value) for value in df[column_name])
        max_length = max((len(value) for value in column_values), default=0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            config.max_column_width,
        )


def build_inspection_workbook_from_csvs(
    workbook_path: Path,
    sheet_specs: list[CsvSheetSpec],
    config: InspectionWorkbookConfig | None = None,
) -> InspectionWorkbookResult:
    """Build a formatted inspection workbook from CSV sources."""

    config = config or InspectionWorkbookConfig()
    workbook = Workbook()
    workbook.remove(workbook.active)

    included_sheets: list[str] = []
    csv_files_used: list[Path] = []
    missing_csv_files: list[Path] = []

    for sheet_spec in sheet_specs:
        if not sheet_spec.csv_path.exists():
            missing_csv_files.append(sheet_spec.csv_path)
            continue

        dataframe = pd.read_csv(sheet_spec.csv_path)
        _write_dataframe_to_sheet(
            workbook=workbook,
            sheet_name=sheet_spec.sheet_name,
            df=dataframe,
            config=config,
        )
        included_sheets.append(sheet_spec.sheet_name)
        csv_files_used.append(sheet_spec.csv_path)

    if not included_sheets:
        raise FileNotFoundError("No CSV files were available for the inspection workbook.")

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(workbook_path)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write inspection workbook because the file is locked or open elsewhere: {workbook_path}"
        ) from exc
    return InspectionWorkbookResult(
        workbook_path=workbook_path,
        included_sheets=tuple(included_sheets),
        csv_files_used=tuple(csv_files_used),
        missing_csv_files=tuple(missing_csv_files),
    )
