from __future__ import annotations

import sys
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.ticker import FuncFormatter, PercentFormatter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from notebooks.utils.inspection_workbooks import (  # noqa: E402
    CsvSheetSpec,
    InspectionWorkbookConfig,
    build_inspection_workbook_from_csvs,
)


PROCESSED_FINANCIALS_RELATIVE_PATH = Path("data/processed/qbts_financials_perplexity_clean.csv")
PROCESSED_FLAGS_RELATIVE_PATH = Path("data/processed/qbts_financials_perplexity_clean_flags.csv")

TABLES_DIR_RELATIVE_PATH = Path("outputs/tables")
CHARTS_DIR_RELATIVE_PATH = Path("outputs/charts")
INSPECTION_WORKBOOK_RELATIVE_PATH = Path("outputs/inspection/final_financial_analysis_inspection.xlsx")
CHART_STYLE_NOTE_RELATIVE_PATH = Path("outputs/text/final_chart_style_note.md")
SUMMARY_RELATIVE_PATH = Path("outputs/text/final_financial_analysis_summary.md")

INCOME_STATEMENT_TABLE_RELATIVE_PATH = TABLES_DIR_RELATIVE_PATH / "final_historical_income_statement_analysis.csv"
MARGIN_TABLE_RELATIVE_PATH = TABLES_DIR_RELATIVE_PATH / "final_historical_margin_analysis.csv"
CASH_FLOW_TABLE_RELATIVE_PATH = TABLES_DIR_RELATIVE_PATH / "final_historical_cash_flow_analysis.csv"
LIQUIDITY_TABLE_RELATIVE_PATH = TABLES_DIR_RELATIVE_PATH / "final_historical_liquidity_analysis.csv"
FLAGS_TABLE_RELATIVE_PATH = TABLES_DIR_RELATIVE_PATH / "final_historical_data_quality_flags_used.csv"

REVENUE_GROWTH_CHART_RELATIVE_PATH = CHARTS_DIR_RELATIVE_PATH / "final_revenue_growth.png"
MARGIN_PROFILE_CHART_RELATIVE_PATH = CHARTS_DIR_RELATIVE_PATH / "final_margin_profile.png"
OPEX_INTENSITY_CHART_RELATIVE_PATH = CHARTS_DIR_RELATIVE_PATH / "final_opex_intensity.png"
CASH_FLOW_CHART_RELATIVE_PATH = CHARTS_DIR_RELATIVE_PATH / "final_cash_flow_and_burn.png"
LIQUIDITY_CHART_RELATIVE_PATH = CHARTS_DIR_RELATIVE_PATH / "final_liquidity_net_cash.png"
SBC_CHART_RELATIVE_PATH = CHARTS_DIR_RELATIVE_PATH / "final_stock_based_compensation.png"

SOURCE_NOTE = "Source: Perplexity financial data, company filings, team analysis."

BLACK = "#000000"
DARK_GRAY = "#404040"
MEDIUM_GRAY = "#808080"
LIGHT_GRAY = "#D9D9D9"
WHITE = "#FFFFFF"

WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
MANUAL_REVIEW_FILL = PatternFill("solid", fgColor="F4CCCC")
INCOMPLETE_FILL = PatternFill("solid", fgColor="D9EAF7")

REQUIRED_FINANCIAL_COLUMNS = {
    "fiscal_year",
    "revenue",
    "cost_of_revenue",
    "gross_profit",
    "r_and_d",
    "sga",
    "operating_income",
    "ebitda",
    "pretax_income",
    "net_income",
    "cash_and_equivalents",
    "marketable_securities",
    "total_debt",
    "operating_cash_flow",
    "capex",
    "free_cash_flow",
    "stock_based_compensation",
    "net_cash",
}

REQUIRED_FLAG_COLUMNS = {
    "fiscal_year",
    "canonical_metric",
    "value_usd_millions",
    "reconciliation_status",
    "data_quality_flag",
    "notes",
}

ANALYSIS_METRICS_USED = {
    "revenue",
    "cost_of_revenue",
    "gross_profit",
    "r_and_d",
    "sga",
    "operating_income",
    "ebitda",
    "pretax_income",
    "net_income",
    "cash_and_equivalents",
    "marketable_securities",
    "total_debt",
    "operating_cash_flow",
    "capex",
    "free_cash_flow",
    "stock_based_compensation",
    "net_cash",
}

USD_COLUMNS = {
    "revenue",
    "cost_of_revenue",
    "gross_profit",
    "r_and_d",
    "sga",
    "operating_income",
    "ebitda",
    "pretax_income",
    "net_income",
    "operating_cash_flow",
    "capex",
    "free_cash_flow",
    "cash_burn",
    "cash_and_equivalents",
    "marketable_securities",
    "total_debt",
    "net_cash",
    "stock_based_compensation",
}

PERCENT_COLUMNS = {
    "revenue_growth",
    "gross_margin",
    "r_and_d_as_pct_revenue",
    "sga_as_pct_revenue",
    "operating_margin",
    "ebitda_margin",
    "pretax_margin",
    "net_margin",
    "operating_cash_flow_margin",
    "free_cash_flow_margin",
    "cash_as_pct_market_cap",
    "stock_based_compensation_as_pct_revenue",
    "stock_based_compensation_as_pct_operating_cash_flow_burn",
}

RATIO_COLUMNS = {
    "debt_to_cash",
    "cash_runway_years",
}


def find_project_root(start: Path) -> Path:
    start = start.resolve()
    for candidate in [start, *start.parents]:
        if (candidate / "data" / "processed").exists():
            return candidate
    raise FileNotFoundError("Could not find the project root containing data/processed.")


def load_inputs(project_root: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    financials_df = pd.read_csv(project_root / PROCESSED_FINANCIALS_RELATIVE_PATH)
    flags_df = pd.read_csv(project_root / PROCESSED_FLAGS_RELATIVE_PATH)
    return financials_df, flags_df


def validate_inputs(financials_df: pd.DataFrame, flags_df: pd.DataFrame) -> None:
    missing_financial_columns = REQUIRED_FINANCIAL_COLUMNS - set(financials_df.columns)
    if missing_financial_columns:
        raise ValueError(f"Processed financials input missing required columns: {sorted(missing_financial_columns)}")

    missing_flag_columns = REQUIRED_FLAG_COLUMNS - set(flags_df.columns)
    if missing_flag_columns:
        raise ValueError(f"Processed flags input missing required columns: {sorted(missing_flag_columns)}")


def safe_divide(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    result = numerator / denominator
    return result.where(denominator.notna() & denominator.ne(0))


def positive_expense(series: pd.Series) -> pd.Series:
    return series.abs()


def build_analysis_base(financials_df: pd.DataFrame) -> pd.DataFrame:
    analysis_df = financials_df.copy().sort_values("fiscal_year").reset_index(drop=True)
    analysis_df["fiscal_year"] = analysis_df["fiscal_year"].astype(int)

    analysis_df["capex"] = analysis_df["capex"].abs()
    analysis_df["free_cash_flow_calculated"] = analysis_df["operating_cash_flow"] - analysis_df["capex"]
    analysis_df["free_cash_flow"] = analysis_df["free_cash_flow_calculated"]
    analysis_df["net_cash_calculated"] = (
        analysis_df["cash_and_equivalents"]
        + analysis_df["marketable_securities"]
        - analysis_df["total_debt"]
    )
    analysis_df["net_cash"] = analysis_df["net_cash_calculated"].where(
        analysis_df[["cash_and_equivalents", "marketable_securities", "total_debt"]].notna().all(axis=1)
    )

    analysis_df["revenue_growth"] = analysis_df["revenue"].pct_change(fill_method=None)
    analysis_df["gross_margin"] = safe_divide(analysis_df["gross_profit"], analysis_df["revenue"])
    analysis_df["r_and_d_as_pct_revenue"] = safe_divide(positive_expense(analysis_df["r_and_d"]), analysis_df["revenue"])
    analysis_df["sga_as_pct_revenue"] = safe_divide(positive_expense(analysis_df["sga"]), analysis_df["revenue"])
    analysis_df["operating_margin"] = safe_divide(analysis_df["operating_income"], analysis_df["revenue"])
    analysis_df["ebitda_margin"] = safe_divide(analysis_df["ebitda"], analysis_df["revenue"])
    analysis_df["pretax_margin"] = safe_divide(analysis_df["pretax_income"], analysis_df["revenue"])
    analysis_df["net_margin"] = safe_divide(analysis_df["net_income"], analysis_df["revenue"])
    analysis_df["operating_cash_flow_margin"] = safe_divide(analysis_df["operating_cash_flow"], analysis_df["revenue"])
    analysis_df["free_cash_flow_margin"] = safe_divide(analysis_df["free_cash_flow"], analysis_df["revenue"])
    analysis_df["cash_burn"] = (-analysis_df["free_cash_flow"]).clip(lower=0)
    analysis_df["cash_runway_years"] = safe_divide(analysis_df["cash_and_equivalents"], analysis_df["cash_burn"])
    analysis_df.loc[analysis_df["cash_burn"].le(0), "cash_runway_years"] = pd.NA
    analysis_df["cash_as_pct_market_cap"] = pd.NA
    analysis_df["debt_to_cash"] = safe_divide(analysis_df["total_debt"], analysis_df["cash_and_equivalents"])
    analysis_df["stock_based_compensation_as_pct_revenue"] = safe_divide(
        analysis_df["stock_based_compensation"],
        analysis_df["revenue"],
    )
    operating_cash_flow_burn = (-analysis_df["operating_cash_flow"]).where(analysis_df["operating_cash_flow"] < 0)
    analysis_df["stock_based_compensation_as_pct_operating_cash_flow_burn"] = safe_divide(
        analysis_df["stock_based_compensation"],
        operating_cash_flow_burn,
    )
    return analysis_df


def build_output_tables(
    analysis_df: pd.DataFrame,
    flags_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    income_statement_analysis_df = analysis_df[
        [
            "fiscal_year",
            "revenue",
            "revenue_growth",
            "cost_of_revenue",
            "gross_profit",
            "r_and_d",
            "sga",
            "operating_income",
            "ebitda",
            "pretax_income",
            "net_income",
        ]
    ].copy()

    margin_analysis_df = analysis_df[
        [
            "fiscal_year",
            "gross_margin",
            "r_and_d_as_pct_revenue",
            "sga_as_pct_revenue",
            "operating_margin",
            "ebitda_margin",
            "pretax_margin",
            "net_margin",
        ]
    ].copy()

    cash_flow_analysis_df = analysis_df[
        [
            "fiscal_year",
            "operating_cash_flow",
            "operating_cash_flow_margin",
            "capex",
            "free_cash_flow",
            "free_cash_flow_margin",
            "cash_burn",
            "stock_based_compensation",
            "stock_based_compensation_as_pct_revenue",
            "stock_based_compensation_as_pct_operating_cash_flow_burn",
        ]
    ].copy()

    liquidity_analysis_df = analysis_df[
        [
            "fiscal_year",
            "cash_and_equivalents",
            "marketable_securities",
            "total_debt",
            "net_cash",
            "cash_runway_years",
            "cash_as_pct_market_cap",
            "debt_to_cash",
        ]
    ].copy()

    data_quality_flags_df = (
        flags_df.loc[flags_df["canonical_metric"].isin(ANALYSIS_METRICS_USED)]
        .sort_values(["fiscal_year", "canonical_metric"])
        .reset_index(drop=True)
    )
    return (
        income_statement_analysis_df,
        margin_analysis_df,
        cash_flow_analysis_df,
        liquidity_analysis_df,
        data_quality_flags_df,
    )


def build_summary_metrics_table(analysis_df: pd.DataFrame, flags_df: pd.DataFrame) -> pd.DataFrame:
    latest = analysis_df.iloc[-1]
    first = analysis_df.iloc[0]
    years_between = int(latest["fiscal_year"] - first["fiscal_year"])
    revenue_cagr = (
        (latest["revenue"] / first["revenue"]) ** (1 / years_between) - 1
        if years_between > 0 and pd.notna(latest["revenue"]) and pd.notna(first["revenue"]) and first["revenue"] > 0
        else pd.NA
    )
    warning_count = int(
        flags_df["data_quality_flag"].eq("annual_used_with_reconciliation_warning").sum()
    )
    manual_review_count = int(flags_df["data_quality_flag"].eq("annual_used_needs_manual_review").sum())
    missing_count = int(flags_df["data_quality_flag"].eq("missing_value").sum())

    rows = [
        {
            "metric": "latest_fiscal_year",
            "value": int(latest["fiscal_year"]),
            "unit": "year",
            "notes": "Latest annual historical period included.",
        },
        {
            "metric": "latest_revenue",
            "value": latest["revenue"],
            "unit": "USD millions",
            "notes": "Latest annual revenue.",
        },
        {
            "metric": "revenue_cagr",
            "value": revenue_cagr,
            "unit": "percentage",
            "notes": f"Compound annual growth from {int(first['fiscal_year'])} to {int(latest['fiscal_year'])}.",
        },
        {
            "metric": "latest_gross_margin",
            "value": latest["gross_margin"],
            "unit": "percentage",
            "notes": "Latest gross margin.",
        },
        {
            "metric": "latest_operating_margin",
            "value": latest["operating_margin"],
            "unit": "percentage",
            "notes": "Latest operating margin.",
        },
        {
            "metric": "latest_free_cash_flow",
            "value": latest["free_cash_flow"],
            "unit": "USD millions",
            "notes": "Operating cash flow less capex.",
        },
        {
            "metric": "latest_cash_burn",
            "value": latest["cash_burn"],
            "unit": "USD millions",
            "notes": "Positive analytical burn measure based on negative free cash flow.",
        },
        {
            "metric": "latest_cash_runway",
            "value": latest["cash_runway_years"],
            "unit": "years",
            "notes": "Cash and equivalents divided by latest annual cash burn; excludes marketable securities.",
        },
        {
            "metric": "latest_net_cash",
            "value": latest["net_cash"],
            "unit": "USD millions",
            "notes": "Calculated only when cash, marketable securities, and debt are all available.",
        },
        {
            "metric": "reconciliation_warning_rows_used",
            "value": warning_count,
            "unit": "count",
            "notes": "Warnings carried into analysis from the processed flags file.",
        },
        {
            "metric": "manual_review_rows_used",
            "value": manual_review_count,
            "unit": "count",
            "notes": "Rows requiring manual review from the processed flags file.",
        },
        {
            "metric": "missing_value_rows_used",
            "value": missing_count,
            "unit": "count",
            "notes": "Missing processed values relevant to the analysis workflow.",
        },
    ]
    return pd.DataFrame(rows, columns=["metric", "value", "unit", "notes"])


def write_tables(
    project_root: Path,
    income_statement_analysis_df: pd.DataFrame,
    margin_analysis_df: pd.DataFrame,
    cash_flow_analysis_df: pd.DataFrame,
    liquidity_analysis_df: pd.DataFrame,
    data_quality_flags_df: pd.DataFrame,
) -> dict[str, Path]:
    output_paths = {
        "income_statement_analysis": project_root / INCOME_STATEMENT_TABLE_RELATIVE_PATH,
        "margin_analysis": project_root / MARGIN_TABLE_RELATIVE_PATH,
        "cash_flow_analysis": project_root / CASH_FLOW_TABLE_RELATIVE_PATH,
        "liquidity_analysis": project_root / LIQUIDITY_TABLE_RELATIVE_PATH,
        "data_quality_flags": project_root / FLAGS_TABLE_RELATIVE_PATH,
    }
    for path in output_paths.values():
        path.parent.mkdir(parents=True, exist_ok=True)

    income_statement_analysis_df.to_csv(output_paths["income_statement_analysis"], index=False)
    margin_analysis_df.to_csv(output_paths["margin_analysis"], index=False)
    cash_flow_analysis_df.to_csv(output_paths["cash_flow_analysis"], index=False)
    liquidity_analysis_df.to_csv(output_paths["liquidity_analysis"], index=False)
    data_quality_flags_df.to_csv(output_paths["data_quality_flags"], index=False)
    return output_paths


def apply_cfa_chart_style(ax: plt.Axes) -> None:
    ax.set_facecolor(WHITE)
    ax.figure.patch.set_facecolor(WHITE)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color(LIGHT_GRAY)
    ax.spines["bottom"].set_color(LIGHT_GRAY)
    ax.grid(axis="y", color=LIGHT_GRAY, linewidth=0.8)
    ax.grid(axis="x", visible=False)
    ax.set_axisbelow(True)
    ax.tick_params(colors=DARK_GRAY, labelsize=9)
    ax.title.set_color(BLACK)
    ax.title.set_fontsize(13)
    ax.title.set_fontweight("bold")
    ax.xaxis.label.set_color(DARK_GRAY)
    ax.yaxis.label.set_color(DARK_GRAY)
    ax.xaxis.label.set_fontsize(10)
    ax.yaxis.label.set_fontsize(10)


def add_source_note(fig: plt.Figure) -> None:
    fig.text(0.01, 0.01, SOURCE_NOTE, ha="left", va="bottom", fontsize=8, color=DARK_GRAY)


def save_chart(fig: plt.Figure, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout(rect=(0, 0.04, 1, 1))
    fig.savefig(output_path, dpi=300, bbox_inches="tight", facecolor=WHITE)
    plt.close(fig)


def format_usd_axis(value: float, _: int) -> str:
    return f"${value:,.0f}"


def create_revenue_growth_chart(analysis_df: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    apply_cfa_chart_style(ax)
    years = analysis_df["fiscal_year"]

    ax.bar(years, analysis_df["revenue"], color=DARK_GRAY, width=0.6, label="Revenue")
    ax.set_title("Revenue Scale and Growth")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("Revenue (USD millions)")
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(FuncFormatter(format_usd_axis))

    growth_df = analysis_df[analysis_df["revenue_growth"].notna()]
    if not growth_df.empty:
        growth_ax = ax.twinx()
        apply_cfa_chart_style(growth_ax)
        growth_ax.spines["left"].set_visible(False)
        growth_ax.grid(False)
        growth_ax.plot(
            growth_df["fiscal_year"],
            growth_df["revenue_growth"],
            color=BLACK,
            linewidth=2.2,
            marker="o",
            label="Revenue growth",
        )
        growth_ax.set_ylabel("Revenue growth")
        growth_ax.yaxis.set_major_formatter(PercentFormatter(1.0))
        handles, labels = [], []
        for axis in (ax, growth_ax):
            axis_handles, axis_labels = axis.get_legend_handles_labels()
            handles.extend(axis_handles)
            labels.extend(axis_labels)
        ax.legend(handles, labels, loc="upper left", frameon=False)
    else:
        ax.legend(loc="upper left", frameon=False)

    add_source_note(fig)
    save_chart(fig, output_path)


def create_margin_profile_chart(analysis_df: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    apply_cfa_chart_style(ax)
    series_specs = [
        ("gross_margin", "Gross margin", BLACK, "-", 2.4),
        ("operating_margin", "Operating margin", DARK_GRAY, "--", 2.0),
        ("ebitda_margin", "EBITDA margin", MEDIUM_GRAY, ":", 2.0),
        ("net_margin", "Net margin", LIGHT_GRAY, "-.", 2.0),
    ]
    for column, label, color, linestyle, linewidth in series_specs:
        if analysis_df[column].notna().any():
            ax.plot(
                analysis_df["fiscal_year"],
                analysis_df[column],
                label=label,
                color=color,
                linestyle=linestyle,
                linewidth=linewidth,
                marker="o",
            )
    ax.set_title("Margin Profile")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("Margin")
    ax.set_xticks(analysis_df["fiscal_year"])
    ax.yaxis.set_major_formatter(PercentFormatter(1.0))
    ax.legend(loc="best", frameon=False)
    add_source_note(fig)
    save_chart(fig, output_path)


def create_opex_intensity_chart(analysis_df: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    apply_cfa_chart_style(ax)
    ax.plot(
        analysis_df["fiscal_year"],
        analysis_df["r_and_d_as_pct_revenue"],
        color=BLACK,
        linewidth=2.4,
        marker="o",
        label="R&D / revenue",
    )
    ax.plot(
        analysis_df["fiscal_year"],
        analysis_df["sga_as_pct_revenue"],
        color=DARK_GRAY,
        linewidth=2.0,
        linestyle="--",
        marker="o",
        label="SG&A / revenue",
    )
    ax.set_title("Operating Expense Intensity")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("Expense intensity")
    ax.set_xticks(analysis_df["fiscal_year"])
    ax.yaxis.set_major_formatter(PercentFormatter(1.0))
    ax.legend(loc="best", frameon=False)
    add_source_note(fig)
    save_chart(fig, output_path)


def create_cash_flow_chart(analysis_df: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    apply_cfa_chart_style(ax)
    years = analysis_df["fiscal_year"]
    width = 0.35
    ax.bar(years - width / 2, analysis_df["operating_cash_flow"], width=width, color=BLACK, label="Operating cash flow")
    ax.bar(years + width / 2, analysis_df["free_cash_flow"], width=width, color=MEDIUM_GRAY, label="Free cash flow")
    ax.axhline(0, color=DARK_GRAY, linewidth=1.0)
    ax.set_title("Cash Flow and Burn")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("USD millions")
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(FuncFormatter(format_usd_axis))
    ax.legend(loc="best", frameon=False)
    add_source_note(fig)
    save_chart(fig, output_path)


def create_liquidity_chart(analysis_df: pd.DataFrame, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    apply_cfa_chart_style(ax)
    years = analysis_df["fiscal_year"]
    width = 0.25
    ax.bar(years - width, analysis_df["cash_and_equivalents"], width=width, color=DARK_GRAY, label="Cash & equivalents")
    ax.bar(years, analysis_df["total_debt"], width=width, color=LIGHT_GRAY, edgecolor=DARK_GRAY, label="Total debt")
    if analysis_df["net_cash"].notna().any():
        ax.bar(years + width, analysis_df["net_cash"], width=width, color=BLACK, label="Net cash")
    ax.set_title("Liquidity and Net Cash Position")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("USD millions")
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(FuncFormatter(format_usd_axis))
    ax.legend(loc="best", frameon=False)
    add_source_note(fig)
    save_chart(fig, output_path)


def create_sbc_chart(analysis_df: pd.DataFrame, output_path: Path) -> bool:
    if not analysis_df["stock_based_compensation"].notna().any():
        return False

    fig, ax = plt.subplots(figsize=(8.5, 5.0))
    apply_cfa_chart_style(ax)
    years = analysis_df["fiscal_year"]
    ax.bar(years, analysis_df["stock_based_compensation"], color=DARK_GRAY, width=0.6, label="SBC")
    ax.set_title("Stock-Based Compensation")
    ax.set_xlabel("Fiscal year")
    ax.set_ylabel("SBC (USD millions)")
    ax.set_xticks(years)
    ax.yaxis.set_major_formatter(FuncFormatter(format_usd_axis))

    ratio_df = analysis_df[analysis_df["stock_based_compensation_as_pct_revenue"].notna()]
    if not ratio_df.empty:
        ratio_ax = ax.twinx()
        apply_cfa_chart_style(ratio_ax)
        ratio_ax.spines["left"].set_visible(False)
        ratio_ax.grid(False)
        ratio_ax.plot(
            ratio_df["fiscal_year"],
            ratio_df["stock_based_compensation_as_pct_revenue"],
            color=BLACK,
            linewidth=2.2,
            marker="o",
            label="SBC / revenue",
        )
        ratio_ax.set_ylabel("SBC / revenue")
        ratio_ax.yaxis.set_major_formatter(PercentFormatter(1.0))
        handles, labels = [], []
        for axis in (ax, ratio_ax):
            axis_handles, axis_labels = axis.get_legend_handles_labels()
            handles.extend(axis_handles)
            labels.extend(axis_labels)
        ax.legend(handles, labels, loc="upper left", frameon=False)
    else:
        ax.legend(loc="upper left", frameon=False)

    add_source_note(fig)
    save_chart(fig, output_path)
    return True


def create_charts(project_root: Path, analysis_df: pd.DataFrame) -> list[Path]:
    chart_paths = [
        project_root / REVENUE_GROWTH_CHART_RELATIVE_PATH,
        project_root / MARGIN_PROFILE_CHART_RELATIVE_PATH,
        project_root / OPEX_INTENSITY_CHART_RELATIVE_PATH,
        project_root / CASH_FLOW_CHART_RELATIVE_PATH,
        project_root / LIQUIDITY_CHART_RELATIVE_PATH,
    ]
    create_revenue_growth_chart(analysis_df, chart_paths[0])
    create_margin_profile_chart(analysis_df, chart_paths[1])
    create_opex_intensity_chart(analysis_df, chart_paths[2])
    create_cash_flow_chart(analysis_df, chart_paths[3])
    create_liquidity_chart(analysis_df, chart_paths[4])

    sbc_path = project_root / SBC_CHART_RELATIVE_PATH
    if create_sbc_chart(analysis_df, sbc_path):
        chart_paths.append(sbc_path)
    return chart_paths


def write_dataframe_sheet(workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=None if pd.isna(value) else value)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        values = [str(column_name)] + ["" if pd.isna(value) else str(value) for value in dataframe.iloc[:, column_index - 1]]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max(len(value) for value in values) + 2, 12),
            40,
        )


def enhance_inspection_workbook(
    workbook_path: Path,
    summary_metrics_df: pd.DataFrame,
) -> None:
    workbook = load_workbook(workbook_path)

    percentage_columns = set(PERCENT_COLUMNS)
    usd_columns = set(USD_COLUMNS)

    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        for column_index, header in enumerate(headers, start=1):
            if header == "fiscal_year":
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column_index).number_format = "0"
            elif header in percentage_columns:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column_index).number_format = "0.0%"
            elif header in usd_columns or header == "value_usd_millions":
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column_index).number_format = '$#,##0.00;[Red]-$#,##0.00'
            elif header in RATIO_COLUMNS:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_index, column=column_index).number_format = "0.00"

    flags_ws = workbook["data_quality_flags"]
    headers = [cell.value for cell in flags_ws[1]]
    quality_flag_column_index = headers.index("data_quality_flag") + 1
    for row_index in range(2, flags_ws.max_row + 1):
        quality_flag = flags_ws.cell(row=row_index, column=quality_flag_column_index).value
        if quality_flag == "annual_used_with_reconciliation_warning":
            fill = WARNING_FILL
        elif quality_flag == "annual_used_needs_manual_review":
            fill = MANUAL_REVIEW_FILL
        elif quality_flag == "annual_used_quarterly_incomplete":
            fill = INCOMPLETE_FILL
        else:
            fill = None
        if fill:
            for cell in flags_ws[row_index]:
                cell.fill = fill

    write_dataframe_sheet(workbook, "summary_metrics", summary_metrics_df)
    summary_ws = workbook["summary_metrics"]
    headers = [cell.value for cell in summary_ws[1]]
    value_column_index = headers.index("value") + 1
    unit_column_index = headers.index("unit") + 1
    for row_index in range(2, summary_ws.max_row + 1):
        unit = summary_ws.cell(row=row_index, column=unit_column_index).value
        value_cell = summary_ws.cell(row=row_index, column=value_column_index)
        if unit == "USD millions":
            value_cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
        elif unit == "percentage":
            value_cell.number_format = "0.0%"
        elif unit == "years":
            value_cell.number_format = "0.00"
        elif unit == "year":
            value_cell.number_format = "0"

    workbook.save(workbook_path)


def build_inspection_workbook(
    project_root: Path,
    table_paths: dict[str, Path],
    summary_metrics_df: pd.DataFrame,
) -> None:
    workbook_path = project_root / INSPECTION_WORKBOOK_RELATIVE_PATH
    workbook_result = build_inspection_workbook_from_csvs(
        workbook_path=workbook_path,
        sheet_specs=[
            CsvSheetSpec("income_statement_analysis", table_paths["income_statement_analysis"]),
            CsvSheetSpec("margin_analysis", table_paths["margin_analysis"]),
            CsvSheetSpec("cash_flow_analysis", table_paths["cash_flow_analysis"]),
            CsvSheetSpec("liquidity_analysis", table_paths["liquidity_analysis"]),
            CsvSheetSpec("data_quality_flags", table_paths["data_quality_flags"]),
        ],
        config=InspectionWorkbookConfig(
            financial_metric_columns=frozenset(USD_COLUMNS),
            explicit_percentage_columns=frozenset(PERCENT_COLUMNS),
            text_label_columns=frozenset({"canonical_metric", "reconciliation_status", "data_quality_flag", "notes"}),
            unmapped_metric_column=None,
        ),
    )
    enhance_inspection_workbook(workbook_path, summary_metrics_df)

    print(f"Created inspection workbook: {workbook_path}")
    print("Included sheets: income_statement_analysis, margin_analysis, cash_flow_analysis, liquidity_analysis, data_quality_flags, summary_metrics")
    print("CSV files used:")
    for csv_file in workbook_result.csv_files_used:
        print(f"- {csv_file}")
    print("CSV files missing:")
    if workbook_result.missing_csv_files:
        for csv_file in workbook_result.missing_csv_files:
            print(f"- {csv_file}")
    else:
        print("- None")


def markdown_table(dataframe: pd.DataFrame) -> str:
    if dataframe.empty:
        return "_None._"
    text_df = dataframe.fillna("").astype(str)
    headers = list(text_df.columns)
    rows = text_df.values.tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(lines)


def format_percentage(value: float | pd.NA) -> str:
    return "N/A" if pd.isna(value) else f"{value:.1%}"


def format_usd(value: float | pd.NA) -> str:
    return "N/A" if pd.isna(value) else f"${value:,.1f} million"


def build_chart_style_note() -> str:
    return "\n".join(
        [
            "# Final chart style note",
            "",
            "## Design intent",
            "",
            "The final historical-analysis charts use a restrained institutional grayscale palette suited to a CFA Research Challenge / equity research report: black, dark gray, medium gray, light gray, and a white background.",
            "",
            "## Common formatting rules",
            "",
            "- Matplotlib only; no seaborn and no bright default color cycle.",
            "- White backgrounds with black or dark-gray text.",
            "- Light-gray horizontal gridlines only, with no heavy chart borders.",
            "- Top and right spines removed; remaining spines kept light gray.",
            "- Clear titles, axis labels, and legends only where they help interpretation.",
            "- Key series use darker or thicker marks; secondary series use lighter grayscale treatment.",
            "- Every chart includes the source note: `Source: Perplexity financial data, company filings, team analysis.`",
            "- Files are saved at 300 dpi with tight layout for report-ready insertion.",
            "",
            "## Reusable implementation",
            "",
            "The script defines `apply_cfa_chart_style(ax)` so the presentation system is consistent across all final historical-analysis charts.",
            "",
        ]
    )


def build_written_summary(
    analysis_df: pd.DataFrame,
    data_quality_flags_df: pd.DataFrame,
) -> str:
    first = analysis_df.iloc[0]
    latest = analysis_df.iloc[-1]
    years = analysis_df["fiscal_year"].astype(int).tolist()
    revenue_cagr = (
        (latest["revenue"] / first["revenue"]) ** (1 / (latest["fiscal_year"] - first["fiscal_year"])) - 1
        if latest["fiscal_year"] > first["fiscal_year"] and first["revenue"] > 0
        else pd.NA
    )
    warning_metrics = sorted(
        data_quality_flags_df.loc[
            data_quality_flags_df["data_quality_flag"] == "annual_used_with_reconciliation_warning",
            "canonical_metric",
        ]
        .drop_duplicates()
        .tolist()
    )
    manual_review_metrics = sorted(
        data_quality_flags_df.loc[
            data_quality_flags_df["data_quality_flag"] == "annual_used_needs_manual_review",
            "canonical_metric",
        ]
        .drop_duplicates()
        .tolist()
    )
    missing_metrics_df = (
        data_quality_flags_df.loc[data_quality_flags_df["data_quality_flag"] == "missing_value", ["canonical_metric", "fiscal_year"]]
        .groupby("canonical_metric", as_index=False)["fiscal_year"]
        .agg(lambda values: ", ".join(str(int(value)) for value in values))
        .rename(columns={"fiscal_year": "years_missing"})
    )
    net_income_divergence_note = (
        "Reported net loss diverged materially from operating loss in the later years, so net income should be interpreted with care because non-operating items can materially affect the bottom line."
    )

    latest_cash_runway = latest["cash_runway_years"]
    latest_sbc_burn_ratio = latest["stock_based_compensation_as_pct_operating_cash_flow_burn"]
    sbc_available = analysis_df["stock_based_compensation"].notna().any()

    lines = [
        "# Final historical financial analysis summary",
        "",
        "## Scope and source policy",
        "",
        f"- Historical years analyzed: {', '.join(str(year) for year in years)}.",
        "- Annual historical data only; no forecasts and no valuation outputs are included in this step.",
        "- All financial values are expressed in USD millions.",
        "- Annual Perplexity financial data remains the authoritative historical source, while the flags file is used only to surface data-quality caveats carried forward from reconciliation.",
        "",
        "## Historical profile",
        "",
        f"Revenue increased from {format_usd(first['revenue'])} in {int(first['fiscal_year'])} to {format_usd(latest['revenue'])} in {int(latest['fiscal_year'])}, equivalent to a historical CAGR of {format_percentage(revenue_cagr)}. The business has grown, but the absolute revenue base remains small for a public-company valuation case.",
        "",
        f"Gross margin moved from {format_percentage(first['gross_margin'])} to {format_percentage(latest['gross_margin'])}. The latest margin profile suggests meaningful gross-profit potential if the company can scale revenue, but the historical record is still too small and uneven to prove durable operating scalability on its own.",
        "",
        f"R&D intensity remained structurally high, reaching {format_percentage(latest['r_and_d_as_pct_revenue'])} of revenue in the latest year. That level is understandable for a quantum / deep-tech company with ongoing technology development, but it also means commercialization must eventually absorb a very large fixed innovation burden.",
        "",
        f"SG&A intensity was {format_percentage(latest['sga_as_pct_revenue'])} in the latest year, underscoring the operating-leverage challenge: even after revenue growth, the corporate cost structure remains far above current sales scale.",
        "",
        f"Operating margin was {format_percentage(latest['operating_margin'])} in {int(latest['fiscal_year'])}, and EBITDA margin was {format_percentage(latest['ebitda_margin'])}. The historical statements therefore still show a substantial distance from breakeven.",
        "",
        f"Net margin reached {format_percentage(latest['net_margin'])} in the latest year. {net_income_divergence_note}",
        "",
        f"Operating cash flow was {format_usd(latest['operating_cash_flow'])} and free cash flow was {format_usd(latest['free_cash_flow'])} in the latest year. Using positive analytical burn convention, annual cash burn was {format_usd(latest['cash_burn'])}.",
        "",
        f"Latest cash and equivalents were {format_usd(latest['cash_and_equivalents'])}; net cash was {format_usd(latest['net_cash'])} where all required components were available. Cash runway based only on cash and equivalents divided by latest annual cash burn was {('N/A' if pd.isna(latest_cash_runway) else f'{latest_cash_runway:.2f} years')}.",
        "",
        "`cash_as_pct_market_cap` was intentionally left blank because market capitalization is not present in the processed historical dataset used for this step.",
        "",
    ]

    if sbc_available:
        lines.extend(
            [
                f"Stock-based compensation was {format_usd(latest['stock_based_compensation'])} in the latest year, equal to {format_percentage(latest['stock_based_compensation_as_pct_revenue'])} of revenue"
                + (
                    f" and {format_percentage(latest_sbc_burn_ratio)} of operating-cash-flow burn."
                    if pd.notna(latest_sbc_burn_ratio)
                    else "."
                ),
                "",
            ]
        )

    lines.extend(
        [
            "## Data-quality flags carried into analysis",
            "",
            "- Reconciliation-warning metrics used in this analysis: "
            + (", ".join(f"`{metric}`" for metric in warning_metrics) if warning_metrics else "None."),
            "- Manual-review metrics used in this analysis: "
            + (", ".join(f"`{metric}`" for metric in manual_review_metrics) if manual_review_metrics else "None."),
            "- Missing processed metrics relevant to this analysis:",
            "",
            markdown_table(missing_metrics_df),
            "",
            "## Interpretation for the investment case",
            "",
            "D-Wave should not be framed as a poor company simply because the historical income statement is immature. The company retains technological optionality. However, the historical fundamentals show limited revenue scale, high operating-expense intensity, negative profitability, and cash burn. A contrarian / fundamentalist valuation approach is therefore appropriate: current valuation must be justified by aggressive future commercialization assumptions rather than by the present historical earnings base alone.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    project_root = find_project_root(Path.cwd())
    financials_df, flags_df = load_inputs(project_root)
    validate_inputs(financials_df, flags_df)

    analysis_df = build_analysis_base(financials_df)
    (
        income_statement_analysis_df,
        margin_analysis_df,
        cash_flow_analysis_df,
        liquidity_analysis_df,
        data_quality_flags_df,
    ) = build_output_tables(analysis_df, flags_df)
    summary_metrics_df = build_summary_metrics_table(analysis_df, data_quality_flags_df)

    table_paths = write_tables(
        project_root=project_root,
        income_statement_analysis_df=income_statement_analysis_df,
        margin_analysis_df=margin_analysis_df,
        cash_flow_analysis_df=cash_flow_analysis_df,
        liquidity_analysis_df=liquidity_analysis_df,
        data_quality_flags_df=data_quality_flags_df,
    )
    build_inspection_workbook(project_root, table_paths, summary_metrics_df)
    chart_paths = create_charts(project_root, analysis_df)

    chart_style_note_path = project_root / CHART_STYLE_NOTE_RELATIVE_PATH
    chart_style_note_path.parent.mkdir(parents=True, exist_ok=True)
    chart_style_note_path.write_text(build_chart_style_note(), encoding="utf-8")

    summary_path = project_root / SUMMARY_RELATIVE_PATH
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(
        build_written_summary(
            analysis_df=analysis_df,
            data_quality_flags_df=data_quality_flags_df,
        ),
        encoding="utf-8",
    )

    print("Saved historical analysis tables:")
    for output_path in table_paths.values():
        print(f"- {output_path}")
    print("Saved historical analysis charts:")
    for chart_path in chart_paths:
        print(f"- {chart_path}")
    print(f"Saved chart design note to: {chart_style_note_path}")
    print(f"Saved historical financial analysis summary to: {summary_path}")


if __name__ == "__main__":
    main()
