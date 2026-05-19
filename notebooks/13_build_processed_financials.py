from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from notebooks.utils.inspection_workbooks import (  # noqa: E402
    CsvSheetSpec,
    InspectionWorkbookConfig,
    build_inspection_workbook_from_csvs,
)


INTERIM_DIR_RELATIVE_PATH = Path("data/interim/perplexity")
PROCESSED_DIR_RELATIVE_PATH = Path("data/processed")
ANNUAL_WIDE_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_annual_wide_standardized.csv"
ANNUAL_LONG_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_annual_long_standardized.csv"
QUARTERLY_WIDE_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_quarterly_wide_standardized.csv"
RECONCILIATION_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_annual_vs_quarterly_reconciliation.csv"
MAPPING_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_line_item_mapping.csv"

PROCESSED_FINANCIALS_RELATIVE_PATH = PROCESSED_DIR_RELATIVE_PATH / "qbts_financials_perplexity_clean.csv"
PROCESSED_FLAGS_RELATIVE_PATH = PROCESSED_DIR_RELATIVE_PATH / "qbts_financials_perplexity_clean_flags.csv"
INSPECTION_WORKBOOK_RELATIVE_PATH = Path("outputs/inspection/qbts_processed_financials_inspection.xlsx")
SUMMARY_RELATIVE_PATH = Path("outputs/text/qbts_processed_financials_summary.md")

REQUIRED_COLUMNS = [
    "fiscal_year",
    "revenue",
    "cost_of_revenue",
    "gross_profit",
    "r_and_d",
    "sga",
    "operating_income",
    "ebitda",
    "depreciation_and_amortization",
    "interest_income",
    "interest_expense",
    "pretax_income",
    "tax_expense",
    "net_income",
    "cash_and_equivalents",
    "marketable_securities",
    "total_debt",
    "total_assets",
    "total_liabilities",
    "stockholders_equity",
    "operating_cash_flow",
    "capex",
    "free_cash_flow",
    "stock_based_compensation",
    "shares_outstanding",
    "weighted_average_basic_shares",
    "weighted_average_diluted_shares",
    "net_cash",
]

DIRECT_ANNUAL_METRICS = {
    "revenue",
    "cost_of_revenue",
    "gross_profit",
    "r_and_d",
    "sga",
    "operating_income",
    "depreciation_and_amortization",
    "interest_income",
    "interest_expense",
    "pretax_income",
    "tax_expense",
    "net_income",
    "cash_and_equivalents",
    "marketable_securities",
    "total_assets",
    "total_liabilities",
    "stockholders_equity",
    "operating_cash_flow",
    "capex",
    "stock_based_compensation",
}

DERIVED_METRICS = {
    "ebitda",
    "total_debt",
    "free_cash_flow",
    "net_cash",
}

SHARE_METRICS = {
    "shares_outstanding",
    "weighted_average_basic_shares",
    "weighted_average_diluted_shares",
}

FINANCIAL_VALUE_COLUMNS = {
    column
    for column in REQUIRED_COLUMNS
    if column not in {"fiscal_year", *SHARE_METRICS}
}

QUALITY_FLAG_BY_RECON_STATUS = {
    "PASS": "annual_used_reconciled",
    "WARNING": "annual_used_with_reconciliation_warning",
    "INCOMPLETE": "annual_used_quarterly_incomplete",
    "FAIL": "annual_used_needs_manual_review",
}

RECON_STATUS_COLORS = {
    "PASS": "C6EFCE",
    "WARNING": "FFF2CC",
    "FAIL": "F4CCCC",
    "INCOMPLETE": "D9EAF7",
    "NOT_FOUND": "E7E6E6",
}

QUALITY_FLAG_COLORS = {
    "annual_used_with_reconciliation_warning": "FFF2CC",
    "annual_used_quarterly_incomplete": "D9EAF7",
    "annual_used_needs_manual_review": "F4CCCC",
    "missing_value": "F4CCCC",
}

IMPORTANT_DISCLOSURE = (
    "Annual Perplexity financial data is used as the authoritative historical source. "
    "Quarterly data is used only as a reconciliation check. "
    "Minor annual-vs-quarterly differences were retained as data-quality warnings rather than adjusted, after manual review."
)


def find_project_root(start: Path) -> Path:
    start = start.resolve()
    for candidate in [start, *start.parents]:
        if (candidate / "data" / "interim" / "perplexity").exists():
            return candidate
    raise FileNotFoundError("Could not find the project root containing data/interim/perplexity.")


def load_inputs(project_root: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    annual_wide_df = pd.read_csv(project_root / ANNUAL_WIDE_RELATIVE_PATH)
    annual_long_df = pd.read_csv(project_root / ANNUAL_LONG_RELATIVE_PATH)
    quarterly_wide_df = pd.read_csv(project_root / QUARTERLY_WIDE_RELATIVE_PATH)
    reconciliation_df = pd.read_csv(project_root / RECONCILIATION_RELATIVE_PATH)
    mapping_df = pd.read_csv(project_root / MAPPING_RELATIVE_PATH)
    return annual_wide_df, annual_long_df, quarterly_wide_df, reconciliation_df, mapping_df


def validate_inputs(
    annual_wide_df: pd.DataFrame,
    annual_long_df: pd.DataFrame,
    quarterly_wide_df: pd.DataFrame,
    reconciliation_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
) -> None:
    required_annual_wide_columns = {
        "fiscal_year",
        "revenue",
        "cost_of_revenue",
        "gross_profit",
        "r_and_d",
        "sga",
        "operating_income",
        "depreciation_and_amortization",
        "interest_income",
        "interest_expense",
        "pretax_income",
        "net_income",
        "cash_and_equivalents",
        "marketable_securities",
        "total_assets",
        "total_liabilities",
        "stockholders_equity",
        "operating_cash_flow",
        "capex",
        "stock_based_compensation",
        "short_term_debt",
        "long_term_debt",
    }
    missing_annual_wide_columns = required_annual_wide_columns - set(annual_wide_df.columns)
    if missing_annual_wide_columns:
        raise ValueError(f"Annual wide input missing required columns: {sorted(missing_annual_wide_columns)}")

    required_annual_long_columns = {
        "fiscal_year",
        "statement_or_tab",
        "canonical_metric",
        "value_usd_millions",
    }
    missing_annual_long_columns = required_annual_long_columns - set(annual_long_df.columns)
    if missing_annual_long_columns:
        raise ValueError(f"Annual long input missing required columns: {sorted(missing_annual_long_columns)}")

    required_quarterly_wide_columns = {"fiscal_year", "fiscal_quarter", "period"}
    missing_quarterly_wide_columns = required_quarterly_wide_columns - set(quarterly_wide_df.columns)
    if missing_quarterly_wide_columns:
        raise ValueError(f"Quarterly wide input missing required columns: {sorted(missing_quarterly_wide_columns)}")

    required_reconciliation_columns = {"fiscal_year", "canonical_metric", "status"}
    missing_reconciliation_columns = required_reconciliation_columns - set(reconciliation_df.columns)
    if missing_reconciliation_columns:
        raise ValueError(f"Reconciliation input missing required columns: {sorted(missing_reconciliation_columns)}")

    required_mapping_columns = {
        "original_line_item",
        "canonical_metric",
        "statement_or_tab",
        "use_in_core_financials",
    }
    missing_mapping_columns = required_mapping_columns - set(mapping_df.columns)
    if missing_mapping_columns:
        raise ValueError(f"Line-item mapping input missing required columns: {sorted(missing_mapping_columns)}")


def is_present(value: object) -> bool:
    return not pd.isna(value)


def select_income_statement_share_metric(
    annual_long_df: pd.DataFrame,
    fiscal_year: int,
    canonical_metric: str,
) -> float | None:
    matches = annual_long_df[
        (annual_long_df["fiscal_year"] == fiscal_year)
        & (annual_long_df["statement_or_tab"] == "Income Statement")
        & (annual_long_df["canonical_metric"] == canonical_metric)
    ]
    if matches.empty:
        return None
    value = matches.iloc[0]["value_usd_millions"]
    return float(value) if is_present(value) else None


def build_processed_financials(
    annual_wide_df: pd.DataFrame,
    annual_long_df: pd.DataFrame,
) -> pd.DataFrame:
    records: list[dict[str, float | int | None]] = []

    for _, row in annual_wide_df.sort_values("fiscal_year").iterrows():
        fiscal_year = int(row["fiscal_year"])
        record: dict[str, float | int | None] = {"fiscal_year": fiscal_year}

        for metric in DIRECT_ANNUAL_METRICS:
            value = row[metric] if metric in annual_wide_df.columns else None
            record[metric] = float(value) if is_present(value) else None

        if is_present(record.get("capex")):
            record["capex"] = abs(float(record["capex"]))

        operating_income = record.get("operating_income")
        depreciation_and_amortization = record.get("depreciation_and_amortization")
        record["ebitda"] = (
            float(operating_income) + float(depreciation_and_amortization)
            if is_present(operating_income) and is_present(depreciation_and_amortization)
            else None
        )

        short_term_debt = row["short_term_debt"] if "short_term_debt" in annual_wide_df.columns else None
        long_term_debt = row["long_term_debt"] if "long_term_debt" in annual_wide_df.columns else None
        record["total_debt"] = (
            float(short_term_debt) + float(long_term_debt)
            if is_present(short_term_debt) and is_present(long_term_debt)
            else None
        )

        operating_cash_flow = record.get("operating_cash_flow")
        capex = record.get("capex")
        record["free_cash_flow"] = (
            float(operating_cash_flow) - float(capex)
            if is_present(operating_cash_flow) and is_present(capex)
            else None
        )

        weighted_average_basic_shares = select_income_statement_share_metric(
            annual_long_df,
            fiscal_year,
            "weighted_average_basic_shares",
        )
        weighted_average_diluted_shares = select_income_statement_share_metric(
            annual_long_df,
            fiscal_year,
            "weighted_average_diluted_shares",
        )
        record["weighted_average_basic_shares"] = weighted_average_basic_shares
        record["weighted_average_diluted_shares"] = weighted_average_diluted_shares
        record["shares_outstanding"] = weighted_average_diluted_shares

        cash = record.get("cash_and_equivalents")
        marketable_securities = record.get("marketable_securities")
        total_debt = record.get("total_debt")
        record["net_cash"] = (
            float(cash) + float(marketable_securities) - float(total_debt)
            if is_present(cash) and is_present(marketable_securities) and is_present(total_debt)
            else None
        )

        records.append(record)

    processed_df = pd.DataFrame(records)
    return processed_df.reindex(columns=REQUIRED_COLUMNS)


def build_reconciliation_lookup(reconciliation_df: pd.DataFrame) -> dict[tuple[int, str], str]:
    return {
        (int(row.fiscal_year), row.canonical_metric): row.status
        for row in reconciliation_df.itertuples(index=False)
    }


def derive_metric_note(metric: str) -> str:
    if metric == "ebitda":
        return "Derived as operating_income + depreciation_and_amortization from annual GAAP inputs."
    if metric == "total_debt":
        return "Derived as short_term_debt + long_term_debt from annual balance-sheet debt line items; total_liabilities was not used."
    if metric == "free_cash_flow":
        return "Derived as operating_cash_flow - capex from annual inputs."
    if metric == "net_cash":
        return "Derived as cash_and_equivalents + marketable_securities - total_debt when all components are available."
    if metric == "shares_outstanding":
        return "Year-end shares_outstanding unavailable in the annual core statement layer; weighted_average_diluted_shares used as the permitted proxy."
    if metric in {"weighted_average_basic_shares", "weighted_average_diluted_shares"}:
        return "Taken from annual Income Statement share-count line item; values are in millions of shares."
    return "Annual authoritative value retained from standardized Perplexity annual data."


def build_flags_file(
    processed_df: pd.DataFrame,
    reconciliation_df: pd.DataFrame,
) -> pd.DataFrame:
    reconciliation_lookup = build_reconciliation_lookup(reconciliation_df)
    rows: list[dict[str, object]] = []

    for _, processed_row in processed_df.iterrows():
        fiscal_year = int(processed_row["fiscal_year"])
        for metric in [column for column in REQUIRED_COLUMNS if column != "fiscal_year"]:
            value = processed_row[metric]
            value_present = is_present(value)
            recon_status = reconciliation_lookup.get((fiscal_year, metric))
            if recon_status is None:
                recon_status = "NOT_FOUND"

            if not value_present:
                data_quality_flag = "missing_value"
                notes = derive_metric_note(metric)
            elif recon_status in QUALITY_FLAG_BY_RECON_STATUS:
                data_quality_flag = QUALITY_FLAG_BY_RECON_STATUS[recon_status]
                notes = derive_metric_note(metric)
            else:
                data_quality_flag = "annual_used_no_reconciliation_test"
                notes = derive_metric_note(metric)

            rows.append(
                {
                    "fiscal_year": fiscal_year,
                    "canonical_metric": metric,
                    "value_usd_millions": float(value) if value_present else None,
                    "reconciliation_status": recon_status,
                    "data_quality_flag": data_quality_flag,
                    "notes": notes,
                }
            )

    return pd.DataFrame(
        rows,
        columns=[
            "fiscal_year",
            "canonical_metric",
            "value_usd_millions",
            "reconciliation_status",
            "data_quality_flag",
            "notes",
        ],
    )


def build_reconciliation_summary_sheet(flags_df: pd.DataFrame) -> pd.DataFrame:
    return (
        flags_df.groupby(["reconciliation_status", "data_quality_flag"], dropna=False)
        .size()
        .reset_index(name="row_count")
        .sort_values(["reconciliation_status", "data_quality_flag"])
        .reset_index(drop=True)
    )


def build_missing_values_summary(processed_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for column in [column for column in processed_df.columns if column != "fiscal_year"]:
        missing_years = processed_df.loc[processed_df[column].isna(), "fiscal_year"].astype(int).tolist()
        if missing_years:
            rows.append(
                {
                    "canonical_metric": column,
                    "missing_count": len(missing_years),
                    "years_missing": ", ".join(str(year) for year in missing_years),
                }
            )
    return pd.DataFrame(rows, columns=["canonical_metric", "missing_count", "years_missing"])


def write_dataframe_sheet(workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        values = [str(column_name)] + ["" if pd.isna(value) else str(value) for value in dataframe.iloc[:, column_index - 1]]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max(len(value) for value in values) + 2, 12),
            36,
        )


def enhance_processed_workbook(
    workbook_path: Path,
    reconciliation_summary_df: pd.DataFrame,
    missing_values_summary_df: pd.DataFrame,
) -> None:
    workbook = load_workbook(workbook_path)

    flags_ws = workbook["data_quality_flags"]
    headers = [cell.value for cell in flags_ws[1]]
    recon_status_idx = headers.index("reconciliation_status") + 1
    data_quality_flag_idx = headers.index("data_quality_flag") + 1

    for row_index in range(2, flags_ws.max_row + 1):
        recon_status = flags_ws.cell(row=row_index, column=recon_status_idx).value
        data_quality_flag = flags_ws.cell(row=row_index, column=data_quality_flag_idx).value
        fill_color = QUALITY_FLAG_COLORS.get(data_quality_flag) or RECON_STATUS_COLORS.get(recon_status)
        if fill_color:
            fill = PatternFill("solid", fgColor=fill_color)
            for cell in flags_ws[row_index]:
                cell.fill = fill

    write_dataframe_sheet(workbook, "reconciliation_summary", reconciliation_summary_df)
    write_dataframe_sheet(workbook, "missing_values_summary", missing_values_summary_df)

    workbook.save(workbook_path)


def markdown_table(dataframe: pd.DataFrame) -> str:
    if dataframe.empty:
        return "_None._"
    text_df = dataframe.fillna("").astype(str)
    headers = list(text_df.columns)
    rows = text_df.values.tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(lines)


def build_written_summary(
    processed_df: pd.DataFrame,
    flags_df: pd.DataFrame,
    reconciliation_df: pd.DataFrame,
    missing_values_summary_df: pd.DataFrame,
) -> str:
    years = processed_df["fiscal_year"].astype(int).tolist()
    warning_metrics = sorted(
        reconciliation_df.loc[reconciliation_df["status"] == "WARNING", "canonical_metric"]
        .drop_duplicates()
        .tolist()
    )
    manual_review_metrics = sorted(
        flags_df.loc[
            flags_df["data_quality_flag"] == "annual_used_needs_manual_review",
            "canonical_metric",
        ]
        .drop_duplicates()
        .tolist()
    )
    missing_metrics = missing_values_summary_df.copy()
    share_proxy_years = flags_df.loc[
        flags_df["canonical_metric"] == "shares_outstanding",
        "fiscal_year",
    ].astype(int).tolist()

    dataset_ready = not manual_review_metrics

    lines = [
        "# QBTS processed historical financials summary",
        "",
        "## Scope",
        "",
        f"- Years included: {', '.join(str(year) for year in years)}.",
        "- Annual standardized Perplexity data was used as the authoritative source for the historical dataset.",
        "- Quarterly Perplexity data was used for reconciliation / quality control only and was not used to replace annual values.",
        "",
        "## Source policy",
        "",
        IMPORTANT_DISCLOSURE,
        "",
        "## Reconciliation warnings retained after review",
        "",
        "- Metrics with reconciliation warnings: "
        + (", ".join(f"`{metric}`" for metric in warning_metrics) if warning_metrics else "None."),
        "- These warning rows were manually reviewed and the annual values were retained, in line with the annual-authoritative policy.",
        "",
        "## Missing metrics",
        "",
        markdown_table(missing_metrics),
        "",
        "## Manual-review items",
        "",
        "- Metrics with FAIL status requiring manual review: "
        + (", ".join(f"`{metric}`" for metric in manual_review_metrics) if manual_review_metrics else "None."),
        "- `shares_outstanding` was not available from the annual core statement layer; `weighted_average_diluted_shares` was used as the permitted proxy for years "
        + ", ".join(str(year) for year in share_proxy_years)
        + ". This should be disclosed and revisited before any per-share valuation work requiring true year-end shares.",
        "",
        "## Derived metrics",
        "",
        "- `ebitda` = `operating_income` + `depreciation_and_amortization`.",
        "- `total_debt` = `short_term_debt` + `long_term_debt`; `total_liabilities` is never used as debt.",
        "- `free_cash_flow` = `operating_cash_flow` - `capex`.",
        "- `net_cash` = `cash_and_equivalents` + `marketable_securities` - `total_debt` only when all components are available.",
        "",
        "## Readiness assessment",
        "",
    ]

    if dataset_ready:
        lines.append(
            "The dataset is ready to proceed into financial analysis, with the stated limitations and data-quality flags preserved for disclosure."
        )
    else:
        lines.append(
            "The dataset should not yet proceed into financial analysis because at least one metric remains flagged for manual review."
        )

    lines.extend(
        [
            "",
            "## Limitations to disclose",
            "",
            "- Some early-year fields remain missing because the annual source did not explicitly provide them; missing values were not invented.",
            "- `tax_expense` is unavailable in the annual core layer for all included years.",
            "- `marketable_securities` is only explicitly available for 2025, so `net_cash` is only calculated when that component is present.",
            "- Minor annual-versus-quarterly differences remain as documented quality warnings rather than adjustments.",
            "- The available share-count fallback is a weighted-average diluted-share proxy, not a true year-end outstanding-share measure.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    project_root = find_project_root(Path.cwd())
    annual_wide_df, annual_long_df, quarterly_wide_df, reconciliation_df, mapping_df = load_inputs(project_root)
    validate_inputs(
        annual_wide_df=annual_wide_df,
        annual_long_df=annual_long_df,
        quarterly_wide_df=quarterly_wide_df,
        reconciliation_df=reconciliation_df,
        mapping_df=mapping_df,
    )

    processed_df = build_processed_financials(
        annual_wide_df=annual_wide_df,
        annual_long_df=annual_long_df,
    )
    flags_df = build_flags_file(
        processed_df=processed_df,
        reconciliation_df=reconciliation_df,
    )
    reconciliation_summary_df = build_reconciliation_summary_sheet(flags_df)
    missing_values_summary_df = build_missing_values_summary(processed_df)

    processed_financials_path = project_root / PROCESSED_FINANCIALS_RELATIVE_PATH
    processed_flags_path = project_root / PROCESSED_FLAGS_RELATIVE_PATH
    workbook_path = project_root / INSPECTION_WORKBOOK_RELATIVE_PATH
    summary_path = project_root / SUMMARY_RELATIVE_PATH

    processed_financials_path.parent.mkdir(parents=True, exist_ok=True)
    processed_df.to_csv(processed_financials_path, index=False)
    flags_df.to_csv(processed_flags_path, index=False)

    workbook_result = build_inspection_workbook_from_csvs(
        workbook_path=workbook_path,
        sheet_specs=[
            CsvSheetSpec(sheet_name="processed_financials", csv_path=processed_financials_path),
            CsvSheetSpec(sheet_name="data_quality_flags", csv_path=processed_flags_path),
        ],
        config=InspectionWorkbookConfig(
            financial_metric_columns=frozenset(FINANCIAL_VALUE_COLUMNS),
            explicit_currency_columns=frozenset({"value_usd_millions"}),
            text_label_columns=frozenset(
                {
                    "canonical_metric",
                    "reconciliation_status",
                    "data_quality_flag",
                    "notes",
                }
            ),
            unmapped_metric_column=None,
        ),
    )
    enhance_processed_workbook(
        workbook_path=workbook_path,
        reconciliation_summary_df=reconciliation_summary_df,
        missing_values_summary_df=missing_values_summary_df,
    )

    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(
        build_written_summary(
            processed_df=processed_df,
            flags_df=flags_df,
            reconciliation_df=reconciliation_df,
            missing_values_summary_df=missing_values_summary_df,
        ),
        encoding="utf-8",
    )

    print(f"Saved processed financials to: {processed_financials_path}")
    print(f"Saved processed financials flags to: {processed_flags_path}")
    print(f"Created processed financials inspection workbook: {workbook_path}")
    print("Included sheets: processed_financials, data_quality_flags, reconciliation_summary, missing_values_summary")
    print("CSV files used:")
    for csv_file in workbook_result.csv_files_used:
        print(f"- {csv_file}")
    print("CSV files missing:")
    if workbook_result.missing_csv_files:
        for csv_file in workbook_result.missing_csv_files:
            print(f"- {csv_file}")
    else:
        print("- None")
    print(f"Saved processed financials summary to: {summary_path}")


if __name__ == "__main__":
    main()
