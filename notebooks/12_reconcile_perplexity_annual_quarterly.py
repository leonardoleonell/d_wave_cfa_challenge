from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from notebooks.utils.inspection_workbooks import (  # noqa: E402
    CsvSheetSpec,
    InspectionWorkbookConfig,
    build_inspection_workbook_from_csvs,
)


INTERIM_DIR_RELATIVE_PATH = Path("data/interim/perplexity")
ANNUAL_LONG_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_annual_long_standardized.csv"
QUARTERLY_LONG_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_quarterly_long_standardized.csv"
ANNUAL_WIDE_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_annual_wide_standardized.csv"
QUARTERLY_WIDE_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_quarterly_wide_standardized.csv"
RECONCILIATION_CSV_RELATIVE_PATH = INTERIM_DIR_RELATIVE_PATH / "qbts_annual_vs_quarterly_reconciliation.csv"
RECONCILIATION_WORKBOOK_RELATIVE_PATH = Path("outputs/inspection/qbts_annual_vs_quarterly_reconciliation.xlsx")
RECONCILIATION_SUMMARY_RELATIVE_PATH = Path("outputs/text/qbts_reconciliation_summary.md")

ABSOLUTE_DIFFERENCE_TOLERANCE = 0.5
PERCENTAGE_DIFFERENCE_TOLERANCE = 0.02

FLOW_METRICS = {
    "revenue",
    "cost_of_revenue",
    "gross_profit",
    "r_and_d",
    "sga",
    "operating_income",
    "ebitda",
    "depreciation_and_amortization",
    "interest_income",
    "interest_expense",
    "pretax_income",
    "tax_expense",
    "net_income",
}

CASH_FLOW_METRICS = {
    "operating_cash_flow",
    "capex",
    "free_cash_flow",
    "stock_based_compensation",
}

BALANCE_SHEET_METRICS = {
    "cash_and_equivalents",
    "marketable_securities",
    "total_current_assets",
    "total_assets",
    "accounts_payable",
    "accrued_expenses",
    "deferred_revenue",
    "short_term_debt",
    "long_term_debt",
    "total_debt",
    "total_current_liabilities",
    "total_liabilities",
    "stockholders_equity",
}

NOT_APPLICABLE_METRICS = {
    "eps_basic",
    "eps_diluted",
    "weighted_average_basic_shares",
    "weighted_average_diluted_shares",
}

RECONCILIATION_COLUMNS = [
    "fiscal_year",
    "canonical_metric",
    "metric_category",
    "reconciliation_method",
    "annual_value_usd_millions",
    "quarterly_comparison_value_usd_millions",
    "difference_usd_millions",
    "difference_percent",
    "quarters_available",
    "quarters_missing",
    "status",
    "notes",
]

STATUS_FILL_COLORS = {
    "PASS": "C6EFCE",
    "WARNING": "FFF2CC",
    "FAIL": "F4CCCC",
    "INCOMPLETE": "D9EAF7",
    "NOT_APPLICABLE": "E7E6E6",
}

FINANCIAL_STATEMENT_TABS = {"Income Statement", "Balance Sheet", "Cash Flow"}


def find_project_root(start: Path) -> Path:
    start = start.resolve()
    for candidate in [start, *start.parents]:
        if (candidate / "data" / "interim" / "perplexity").exists():
            return candidate
    raise FileNotFoundError("Could not find the project root containing data/interim/perplexity.")


def load_inputs(project_root: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    annual_long_df = pd.read_csv(project_root / ANNUAL_LONG_RELATIVE_PATH)
    quarterly_long_df = pd.read_csv(project_root / QUARTERLY_LONG_RELATIVE_PATH)
    annual_wide_df = pd.read_csv(project_root / ANNUAL_WIDE_RELATIVE_PATH)
    quarterly_wide_df = pd.read_csv(project_root / QUARTERLY_WIDE_RELATIVE_PATH)
    return annual_long_df, quarterly_long_df, annual_wide_df, quarterly_wide_df


def is_present(value: object) -> bool:
    return not pd.isna(value)


def metric_category(metric: str) -> str:
    if metric in FLOW_METRICS:
        return "income_statement_flow"
    if metric in CASH_FLOW_METRICS:
        return "cash_flow"
    if metric in BALANCE_SHEET_METRICS:
        return "balance_sheet_point_in_time"
    return "not_applicable"


def reconciliation_method(category: str) -> str:
    if category in {"income_statement_flow", "cash_flow"}:
        return "annual_value_vs_sum_of_q1_to_q4"
    if category == "balance_sheet_point_in_time":
        return "annual_year_end_value_vs_q4_ending_balance"
    return "not_applicable"


def difference_percent(annual_value: float | None, difference_value: float | None) -> float | None:
    if annual_value is None or difference_value is None:
        return None
    if pd.isna(annual_value) or pd.isna(difference_value) or annual_value == 0:
        return None
    return abs(difference_value) / abs(annual_value)


def assess_status(
    annual_value: float | None,
    quarterly_value: float | None,
    diff_value: float | None,
    diff_pct: float | None,
    quarters_missing: list[str],
    category: str,
) -> str:
    if category == "not_applicable":
        return "NOT_APPLICABLE"
    if annual_value is None or pd.isna(annual_value):
        return "INCOMPLETE"
    if quarters_missing:
        return "INCOMPLETE"
    if quarterly_value is None or pd.isna(quarterly_value):
        return "INCOMPLETE"
    if diff_value is None or pd.isna(diff_value):
        return "INCOMPLETE"

    abs_pass = abs(diff_value) <= ABSOLUTE_DIFFERENCE_TOLERANCE
    pct_pass = diff_pct is not None and diff_pct <= PERCENTAGE_DIFFERENCE_TOLERANCE
    if abs_pass or pct_pass:
        return "PASS"
    if abs(diff_value) <= ABSOLUTE_DIFFERENCE_TOLERANCE * 2 or (
        diff_pct is not None and diff_pct <= PERCENTAGE_DIFFERENCE_TOLERANCE * 2
    ):
        return "WARNING"
    return "FAIL"


def build_notes(
    category: str,
    annual_value: float | None,
    quarters_missing: list[str],
    diff_pct: float | None,
) -> str:
    if category == "not_applicable":
        return "Metric is not economically valid for annual-vs-quarterly reconciliation."
    if annual_value is None or pd.isna(annual_value):
        return "Annual value unavailable in the interim core dataset."
    if quarters_missing:
        return "Missing required quarterly observations: " + ", ".join(quarters_missing) + "."
    if diff_pct is None and annual_value == 0:
        return "Annual value is zero; percentage difference is not meaningful, so status relies on the absolute tolerance."
    if category == "balance_sheet_point_in_time":
        return "Compared annual year-end balance with Q4 ending balance."
    return "Compared annual value with the sum of Q1 through Q4."


def build_reconciliation_table(
    annual_long_df: pd.DataFrame,
    annual_wide_df: pd.DataFrame,
    quarterly_wide_df: pd.DataFrame,
) -> pd.DataFrame:
    annual_years = sorted(int(year) for year in annual_wide_df["fiscal_year"].dropna().unique())
    wide_metrics = {
        column
        for column in annual_wide_df.columns
        if column != "fiscal_year"
    }
    applicable_long_df = annual_long_df[
        annual_long_df["statement_or_tab"].isin(FINANCIAL_STATEMENT_TABS)
    ]
    standardized_metrics = {
        metric
        for metric in applicable_long_df["canonical_metric"].dropna().unique()
        if metric != "UNMAPPED"
    }
    reconciliation_universe = sorted(
        (wide_metrics & (FLOW_METRICS | CASH_FLOW_METRICS | BALANCE_SHEET_METRICS))
        | (standardized_metrics & NOT_APPLICABLE_METRICS)
    )
    annual_long_metric_availability = (
        annual_long_df.loc[
            annual_long_df["statement_or_tab"].isin(FINANCIAL_STATEMENT_TABS)
            & annual_long_df["canonical_metric"].isin(FLOW_METRICS | CASH_FLOW_METRICS | BALANCE_SHEET_METRICS),
            ["fiscal_year", "canonical_metric"],
        ]
        .drop_duplicates()
        .assign(annual_source_metric_available=True)
    )

    quarterly_lookup = quarterly_wide_df.copy()
    records: list[dict[str, object]] = []

    for fiscal_year in annual_years:
        annual_row = annual_wide_df.loc[annual_wide_df["fiscal_year"] == fiscal_year]
        if annual_row.empty:
            continue
        annual_row = annual_row.iloc[0]

        quarterly_year_df = quarterly_lookup.loc[quarterly_lookup["fiscal_year"] == fiscal_year].copy()

        for metric in reconciliation_universe:
            category = metric_category(metric)
            method = reconciliation_method(category)
            annual_value = annual_row[metric] if metric in annual_wide_df.columns else None
            annual_source_metric_available = not annual_long_metric_availability.loc[
                (annual_long_metric_availability["fiscal_year"] == fiscal_year)
                & (annual_long_metric_availability["canonical_metric"] == metric)
            ].empty

            if category in {"income_statement_flow", "cash_flow"}:
                quarterly_values = {}
                for quarter in range(1, 5):
                    quarter_row = quarterly_year_df.loc[quarterly_year_df["fiscal_quarter"] == quarter]
                    quarter_value = None
                    if not quarter_row.empty and metric in quarterly_year_df.columns:
                        candidate = quarter_row.iloc[0][metric]
                        quarter_value = candidate if is_present(candidate) else None
                    quarterly_values[f"Q{quarter}"] = quarter_value
                quarters_available = [quarter for quarter, value in quarterly_values.items() if is_present(value)]
                quarters_missing = [quarter for quarter, value in quarterly_values.items() if not is_present(value)]
                quarterly_value = (
                    sum(float(value) for value in quarterly_values.values() if is_present(value))
                    if not quarters_missing
                    else None
                )
            elif category == "balance_sheet_point_in_time":
                q4_row = quarterly_year_df.loc[quarterly_year_df["fiscal_quarter"] == 4]
                q4_value = None
                if not q4_row.empty and metric in quarterly_year_df.columns:
                    candidate = q4_row.iloc[0][metric]
                    q4_value = candidate if is_present(candidate) else None
                quarterly_values = {"Q4": q4_value}
                quarters_available = [quarter for quarter, value in quarterly_values.items() if is_present(value)]
                quarters_missing = [quarter for quarter, value in quarterly_values.items() if not is_present(value)]
                quarterly_value = float(q4_value) if is_present(q4_value) else None
            else:
                quarters_available = []
                quarters_missing = []
                quarterly_value = None

            if (
                category != "not_applicable"
                and not annual_source_metric_available
                and not is_present(annual_value)
                and not quarters_available
            ):
                continue

            diff_value = (
                float(annual_value) - float(quarterly_value)
                if is_present(annual_value) and is_present(quarterly_value)
                else None
            )
            diff_pct = difference_percent(
                float(annual_value) if is_present(annual_value) else None,
                diff_value,
            )
            status = assess_status(
                annual_value=float(annual_value) if is_present(annual_value) else None,
                quarterly_value=quarterly_value,
                diff_value=diff_value,
                diff_pct=diff_pct,
                quarters_missing=quarters_missing,
                category=category,
            )
            notes = build_notes(
                category=category,
                annual_value=float(annual_value) if is_present(annual_value) else None,
                quarters_missing=quarters_missing,
                diff_pct=diff_pct,
            )

            records.append(
                {
                    "fiscal_year": fiscal_year,
                    "canonical_metric": metric,
                    "metric_category": category,
                    "reconciliation_method": method,
                    "annual_value_usd_millions": float(annual_value) if is_present(annual_value) else None,
                    "quarterly_comparison_value_usd_millions": quarterly_value,
                    "difference_usd_millions": diff_value,
                    "difference_percent": diff_pct,
                    "quarters_available": ", ".join(quarters_available),
                    "quarters_missing": ", ".join(quarters_missing),
                    "status": status,
                    "notes": notes,
                }
            )

    reconciliation_df = pd.DataFrame(records, columns=RECONCILIATION_COLUMNS)
    return reconciliation_df.sort_values(["fiscal_year", "metric_category", "canonical_metric"]).reset_index(drop=True)


def build_summary_table(reconciliation_df: pd.DataFrame) -> pd.DataFrame:
    return (
        reconciliation_df.groupby(["status", "metric_category"], dropna=False)
        .size()
        .reset_index(name="row_count")
        .sort_values(["status", "metric_category"])
        .reset_index(drop=True)
    )


def enhance_reconciliation_workbook(workbook_path: Path, summary_df: pd.DataFrame) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path)
    reconciliation_ws = workbook["reconciliation"]

    header_values = [cell.value for cell in reconciliation_ws[1]]
    status_col_idx = header_values.index("status") + 1
    difference_percent_col_idx = header_values.index("difference_percent") + 1
    currency_columns = {
        "annual_value_usd_millions",
        "quarterly_comparison_value_usd_millions",
        "difference_usd_millions",
    }

    for row_idx in range(2, reconciliation_ws.max_row + 1):
        status = reconciliation_ws.cell(row=row_idx, column=status_col_idx).value
        fill_color = STATUS_FILL_COLORS.get(status)
        if fill_color:
            fill = PatternFill("solid", fgColor=fill_color)
            for cell in reconciliation_ws[row_idx]:
                cell.fill = fill
        reconciliation_ws.cell(row=row_idx, column=difference_percent_col_idx).number_format = "0.0%"

    for col_idx, header in enumerate(header_values, start=1):
        if header in currency_columns:
            for row_idx in range(2, reconciliation_ws.max_row + 1):
                reconciliation_ws.cell(row=row_idx, column=col_idx).number_format = '$#,##0.00;[Red]-$#,##0.00'

    summary_ws = workbook.create_sheet(title="summary_by_status_category")
    for col_idx, header in enumerate(summary_df.columns, start=1):
        cell = summary_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row_idx, row in enumerate(summary_df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                fill_color = STATUS_FILL_COLORS.get(value)
                if fill_color:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
    summary_ws.freeze_panes = "A2"
    summary_ws.auto_filter.ref = summary_ws.dimensions
    for col_idx, column_name in enumerate(summary_df.columns, start=1):
        values = [str(column_name)] + [str(value) for value in summary_df.iloc[:, col_idx - 1].tolist()]
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max(len(value) for value in values) + 2, 12), 28)

    workbook.save(workbook_path)


def build_markdown_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "_None._"
    text_df = df.fillna("").astype(str)
    headers = list(text_df.columns)
    rows = text_df.values.tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(lines)


def build_reconciliation_summary(reconciliation_df: pd.DataFrame) -> str:
    years_reconciled = sorted(reconciliation_df["fiscal_year"].dropna().unique().tolist())
    applicable_df = reconciliation_df[reconciliation_df["status"] != "NOT_APPLICABLE"]
    status_counts = build_summary_table(reconciliation_df)
    failed_df = reconciliation_df[reconciliation_df["status"] == "FAIL"][
        ["fiscal_year", "canonical_metric", "difference_usd_millions", "difference_percent"]
    ]
    incomplete_df = reconciliation_df[reconciliation_df["status"] == "INCOMPLETE"][
        ["fiscal_year", "canonical_metric", "quarters_missing"]
    ]
    warning_df = reconciliation_df[reconciliation_df["status"] == "WARNING"][
        ["fiscal_year", "canonical_metric", "difference_usd_millions", "difference_percent"]
    ]

    complete_applicable_df = applicable_df[applicable_df["status"].isin({"PASS", "WARNING", "FAIL"})]
    annual_data_reliable = (
        not complete_applicable_df.empty
        and (complete_applicable_df["status"] == "FAIL").sum() == 0
    )

    manual_review_metrics = sorted(
        set(failed_df["canonical_metric"].tolist())
        | set(warning_df["canonical_metric"].tolist())
        | set(incomplete_df["canonical_metric"].tolist())
    )

    lines = [
        "# QBTS annual-vs-quarterly reconciliation summary",
        "",
        "## Scope",
        "",
        f"- Years evaluated: {', '.join(str(year) for year in years_reconciled)}.",
        "- Flow metrics are reconciled as annual value versus the sum of Q1 through Q4.",
        "- Balance-sheet metrics are reconciled as annual year-end value versus Q4 ending balance.",
        f"- Tolerances: absolute difference ≤ {ABSOLUTE_DIFFERENCE_TOLERANCE:.1f} USD million or percentage difference ≤ {PERCENTAGE_DIFFERENCE_TOLERANCE:.1%}.",
        "",
        "## Status summary",
        "",
        build_markdown_table(status_counts),
        "",
        "## Metrics that failed",
        "",
        build_markdown_table(failed_df),
        "",
        "## Metrics with warnings",
        "",
        build_markdown_table(warning_df),
        "",
        "## Metrics incomplete because quarters are missing",
        "",
        build_markdown_table(incomplete_df),
        "",
        "## Readiness assessment",
        "",
    ]

    if annual_data_reliable:
        lines.append(
            "For metric-year combinations with complete quarterly support, the annual interim data appears reliable enough to proceed toward final processed-dataset design, subject to manual review of any WARNING rows and the still-incomplete years."
        )
    else:
        lines.append(
            "The annual interim data is **not yet ready** to be promoted into a final processed dataset without review because at least one complete metric-year reconciliation failed."
        )

    lines.extend(
        [
            "",
            "## Metrics to review before final processing",
            "",
            "- Metrics requiring manual review: "
            + (", ".join(f"`{metric}`" for metric in manual_review_metrics) if manual_review_metrics else "None."),
            "- Metrics marked `INCOMPLETE` are not evidence of inconsistency; they indicate missing required quarterly observations.",
            "- Metrics marked `NOT_APPLICABLE` are intentionally excluded from sum-based reconciliation because the economics do not support that comparison.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    project_root = find_project_root(Path.cwd())
    annual_long_df, quarterly_long_df, annual_wide_df, quarterly_wide_df = load_inputs(project_root)

    reconciliation_df = build_reconciliation_table(
        annual_long_df=annual_long_df,
        annual_wide_df=annual_wide_df,
        quarterly_wide_df=quarterly_wide_df,
    )
    summary_df = build_summary_table(reconciliation_df)

    reconciliation_csv_path = project_root / RECONCILIATION_CSV_RELATIVE_PATH
    reconciliation_workbook_path = project_root / RECONCILIATION_WORKBOOK_RELATIVE_PATH
    reconciliation_summary_path = project_root / RECONCILIATION_SUMMARY_RELATIVE_PATH

    reconciliation_csv_path.parent.mkdir(parents=True, exist_ok=True)
    reconciliation_df.to_csv(reconciliation_csv_path, index=False)

    workbook_result = build_inspection_workbook_from_csvs(
        workbook_path=reconciliation_workbook_path,
        sheet_specs=[CsvSheetSpec(sheet_name="reconciliation", csv_path=reconciliation_csv_path)],
        config=InspectionWorkbookConfig(
            financial_metric_columns=frozenset(),
            explicit_currency_columns=frozenset(
                {
                    "annual_value_usd_millions",
                    "quarterly_comparison_value_usd_millions",
                    "difference_usd_millions",
                }
            ),
            explicit_percentage_columns=frozenset({"difference_percent"}),
            text_label_columns=frozenset(
                {
                    "canonical_metric",
                    "metric_category",
                    "reconciliation_method",
                    "quarters_available",
                    "quarters_missing",
                    "status",
                    "notes",
                }
            ),
            unmapped_metric_column=None,
        ),
    )
    enhance_reconciliation_workbook(reconciliation_workbook_path, summary_df)

    reconciliation_summary_path.parent.mkdir(parents=True, exist_ok=True)
    reconciliation_summary_path.write_text(
        build_reconciliation_summary(reconciliation_df),
        encoding="utf-8",
    )

    print(f"Saved reconciliation CSV to: {reconciliation_csv_path}")
    print(f"Created reconciliation inspection workbook: {reconciliation_workbook_path}")
    print("Included sheets: reconciliation, summary_by_status_category")
    print("CSV files used:")
    for csv_file in workbook_result.csv_files_used:
        print(f"- {csv_file}")
    print("CSV files missing:")
    if workbook_result.missing_csv_files:
        for csv_file in workbook_result.missing_csv_files:
            print(f"- {csv_file}")
    else:
        print("- None")
    print(f"Saved reconciliation summary to: {reconciliation_summary_path}")


if __name__ == "__main__":
    main()
